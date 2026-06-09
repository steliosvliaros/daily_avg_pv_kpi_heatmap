from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
SILVER_CSV = ROOT / "outputs" / "silver_daily_energy_2015_to_date.csv"
PVGIS_CSV = ROOT / "outputs" / "pvgis_typical_daily_export.csv"
PARK_METADATA_CSV = ROOT / "mappings" / "park_metadata.csv"
OUT_XLSX = ROOT / "outputs" / "actual_vs_pvgis_workbook.xlsx"


def _to_python_in_excel_formula(code: str) -> str:
    return '=PY("' + code.replace('"', '""') + '")'


def _read_silver_daily_energy(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", dtype={"park_id": "string"}, low_memory=False)

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    if "day_of_year" not in df.columns:
        df["day_of_year"] = df["date"].dt.dayofyear

    # Silver export can have decimal commas depending on locale.
    value_as_text = df["value"].astype("string").str.replace(",", ".", regex=False)
    df["value"] = pd.to_numeric(value_as_text, errors="coerce")

    return df


def _read_pvgis(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", dtype={"park_id": "string"}, low_memory=False)
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df["day_of_year"] = pd.to_numeric(df["day_of_year"], errors="coerce").astype("Int64")
    return df


def _read_park_metadata(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, dtype={"park_id": "string"}, low_memory=False)
    if "park_id" in df.columns:
        df["park_id"] = df["park_id"].astype("string").str.strip().str.lower()

    if "price_euro_to_kwh" in df.columns:
        price_as_text = df["price_euro_to_kwh"].astype("string").str.replace(",", ".", regex=False)
        df["price_euro_to_kwh"] = pd.to_numeric(price_as_text, errors="coerce")

    return df


def _build_range_detail_data(
    daily_out: pd.DataFrame,
    pvgis_out: pd.DataFrame,
    parks: pd.DataFrame,
) -> pd.DataFrame:
    min_date = daily_out["date"].min()
    max_date = daily_out["date"].max()
    if pd.isna(min_date) or pd.isna(max_date) or parks.empty:
        return pd.DataFrame(
            columns=[
                "date",
                "day_of_year",
                "park_id",
                "park_name",
                "actual_value",
                "pvgis_value",
                "actual_over_pvgis",
            ]
        )

    dates = pd.DataFrame({"date": pd.date_range(min_date, max_date, freq="D")})
    grid = dates.assign(_key=1).merge(
        parks[["park_id", "park_name"]].assign(_key=1),
        on="_key",
        how="inner",
    ).drop(columns="_key")
    grid["day_of_year"] = grid["date"].dt.dayofyear

    actual_daily = (
        daily_out.groupby(["date", "park_id"], as_index=False)["value"]
        .sum()
        .rename(columns={"value": "actual_value"})
    )
    pvgis_daily = (
        pvgis_out.groupby(["park_id", "day_of_year"], as_index=False)["value"]
        .sum()
        .rename(columns={"value": "pvgis_value"})
    )

    detail = grid.merge(actual_daily, on=["date", "park_id"], how="left")
    detail = detail.merge(pvgis_daily, on=["park_id", "day_of_year"], how="left")
    detail["actual_value"] = detail["actual_value"].fillna(0)
    detail["pvgis_value"] = detail["pvgis_value"].fillna(0)
    detail["actual_over_pvgis"] = detail["actual_value"] / detail["pvgis_value"].where(detail["pvgis_value"] != 0)

    return detail[
        [
            "date",
            "day_of_year",
            "park_id",
            "park_name",
            "actual_value",
            "pvgis_value",
            "actual_over_pvgis",
        ]
    ].sort_values(["date", "park_name", "park_id"]).reset_index(drop=True)


def _write_helper_daily_analysis_sheet(ws, daily_last_row: int, pvgis_last_row: int) -> None:
    headers = [
        "park_name",
        "sensor_name",
        "park_iso_name",
        "park_id",
        "date",
        "actual_value",
        "day_of_year",
        "pvgis_value",
        "actual_over_pvgis",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    source_cols = ["A", "B", "C", "D", "E", "F", "G"]
    for row_idx in range(2, daily_last_row + 1):
        for col_idx, source_col in enumerate(source_cols, start=1):
            ws.cell(row=row_idx, column=col_idx).value = f"='Daily Energy'!{source_col}{row_idx}"
        ws.cell(row=row_idx, column=8).value = (
            f"=SUMIFS('PVGIS Data'!$D$2:$D${pvgis_last_row},"
            f"'PVGIS Data'!$C$2:$C${pvgis_last_row},D{row_idx},"
            f"'PVGIS Data'!$E$2:$E${pvgis_last_row},G{row_idx})"
        )
        ws.cell(row=row_idx, column=9).value = f"=IFERROR(F{row_idx}/H{row_idx},NA())"
        ws.cell(row=row_idx, column=5).number_format = "yyyy-mm-dd"

    for col, w in [("A", 42), ("B", 28), ("C", 28), ("D", 34), ("E", 12), ("F", 16), ("G", 12), ("H", 16), ("I", 18)]:
        ws.column_dimensions[col].width = w


def _build_python_in_excel_heatmap_formula(
    data_range: str,
    from_date_cell: str,
    to_date_cell: str,
    metric_cell: str,
    title_prefix: str,
) -> str:
    code = (
        "import matplotlib\n"
        "matplotlib.use('Agg')\n"
        "import matplotlib.pyplot as plt\n"
        "import numpy as np\n"
        "import pandas as pd\n"
        f"data = xl('{data_range}', headers=True).copy()\n"
        "data.columns = [str(col).strip() for col in data.columns]\n"
        f"metric_selection = xl('{metric_cell}')\n"
        "if metric_selection is None:\n"
        "    metric_selection = 'Actual/PVGIS'\n"
        "metric_selection = str(metric_selection).strip().lower().replace('_', ' ')\n"
        "metric_map = {\n"
        "    'actual': ('actual_value_range', 'Actual'),\n"
        "    'pvgis': ('pvgis_value_range', 'PVGIS'),\n"
        "    'actual/pvgis': ('actual_over_pvgis', 'Actual/PVGIS'),\n"
        "    'actual over pvgis': ('actual_over_pvgis', 'Actual/PVGIS'),\n"
        "}\n"
        "if metric_selection not in metric_map:\n"
        "    raise ValueError('Metric to plot must be Actual, PVGIS, or Actual/PVGIS')\n"
        "metric_col, metric_label = metric_map[metric_selection]\n"
        "required_cols = ['park_name', metric_col]\n"
        "missing_cols = [col for col in required_cols if col not in data.columns]\n"
        "if missing_cols:\n"
        "    raise ValueError(f'Missing required heatmap columns: {missing_cols}')\n"
        "data[metric_col] = pd.to_numeric(data[metric_col], errors='coerce')\n"
        "data = data.dropna(subset=['park_name'])\n"
        "data = data.dropna(subset=[metric_col])\n"
        "if data.empty:\n"
        "    raise ValueError('No data available for heatmap rendering')\n"
        "display_df = data.set_index('park_name')[[metric_col]].copy()\n"
        "display_df.index = display_df.index.astype(str).str.slice(0, 42)\n"
        "display_df.columns = [metric_label]\n"
        "values = display_df[metric_label].astype(float)\n"
        "if metric_col == 'actual_over_pvgis':\n"
        "    plot_values = values.clip(lower=0, upper=2) / 2.0\n"
        "    labels = values.map(lambda x: '' if pd.isna(x) else f'{x:.2f}x')\n"
        "    cbar_label = 'Actual/PVGIS intensity, clipped at 2x'\n"
        "else:\n"
        "    max_val = values.max(skipna=True)\n"
        "    plot_values = values / max_val if pd.notna(max_val) and max_val > 0 else values\n"
        "    labels = values.map(lambda x: '' if pd.isna(x) else f'{x:,.0f}')\n"
        "    cbar_label = 'Relative intensity'\n"
        "plot_df = pd.DataFrame({metric_label: plot_values}, index=display_df.index)\n"
        "fig, ax = plt.subplots(figsize=(8, max(6, 0.28 * len(plot_df.index) + 1.5)))\n"
        "im = ax.imshow(plot_df.values, aspect='auto', interpolation='nearest', cmap='turbo', vmin=0, vmax=1)\n"
        f"from_date = xl('{from_date_cell}')\n"
        f"to_date = xl('{to_date_cell}')\n"
        f"title = '{title_prefix} - ' + metric_label\n"
        "if from_date is not None and to_date is not None:\n"
        "    title = f'{title} | {from_date:%Y-%m-%d} to {to_date:%Y-%m-%d}'\n"
        "ax.set_title(title, fontsize=14, fontweight='bold')\n"
        "ax.set_yticks(np.arange(len(plot_df.index)))\n"
        "ax.set_yticklabels(plot_df.index, fontsize=10)\n"
        "ax.set_xticks(np.arange(len(plot_df.columns)))\n"
        "ax.set_xticklabels([metric_label], rotation=0, ha='center', fontsize=10)\n"
        "for row_idx, park_name in enumerate(display_df.index):\n"
        "    label = labels.iloc[row_idx]\n"
        "    if not label:\n"
        "        continue\n"
        "    ax.text(0, row_idx, label, ha='center', va='center', fontsize=8, color='white' if plot_df.iloc[row_idx, 0] < 0.35 else 'black')\n"
        "cbar = plt.colorbar(im, ax=ax, fraction=0.025, pad=0.02)\n"
        "cbar.ax.set_ylabel(cbar_label, rotation=90, fontsize=10)\n"
        "ax.grid(False)\n"
        "plt.tight_layout()\n"
        "fig"
    )
    return _to_python_in_excel_formula(code)


def _build_python_in_excel_daily_heatmap_formula(
    data_range: str,
    from_date_cell: str,
    to_date_cell: str,
    metric_cell: str,
    title_prefix: str,
) -> str:
    code = (
        "import matplotlib\n"
        "matplotlib.use('Agg')\n"
        "import matplotlib.pyplot as plt\n"
        "import numpy as np\n"
        "import pandas as pd\n"
        f"data = xl({data_range!r}, headers=True).copy()\n"
        "data.columns = [str(col).strip() for col in data.columns]\n"
        "data['date'] = pd.to_datetime(data['date'], errors='coerce')\n"
        f"from_date = pd.to_datetime(xl('{from_date_cell}'), errors='coerce')\n"
        f"to_date = pd.to_datetime(xl('{to_date_cell}'), errors='coerce')\n"
        f"metric_selection = xl('{metric_cell}')\n"
        "if metric_selection is None:\n"
        "    metric_selection = 'Actual/PVGIS'\n"
        "metric_selection = str(metric_selection).strip().lower().replace('_', ' ')\n"
        "metric_map = {\n"
        "    'actual': ('actual_value', 'Actual'),\n"
        "    'pvgis': ('pvgis_value', 'PVGIS'),\n"
        "    'actual/pvgis': ('actual_over_pvgis', 'Actual/PVGIS'),\n"
        "    'actual over pvgis': ('actual_over_pvgis', 'Actual/PVGIS'),\n"
        "}\n"
        "if metric_selection not in metric_map:\n"
        "    raise ValueError('Metric to plot must be Actual, PVGIS, or Actual/PVGIS')\n"
        "metric_col, metric_label = metric_map[metric_selection]\n"
        "required_cols = ['date', 'park_name', metric_col]\n"
        "missing_cols = [col for col in required_cols if col not in data.columns]\n"
        "if missing_cols:\n"
        "    raise ValueError(f'Missing required heatmap columns: {missing_cols}')\n"
        "if pd.notna(from_date):\n"
        "    data = data[data['date'] >= from_date]\n"
        "if pd.notna(to_date):\n"
        "    data = data[data['date'] <= to_date]\n"
        "data[metric_col] = pd.to_numeric(data[metric_col], errors='coerce')\n"
        "data = data.dropna(subset=['date', 'park_name', metric_col])\n"
        "if data.empty:\n"
        "    raise ValueError('No data available for heatmap rendering')\n"
        "data['date_label'] = data['date'].dt.strftime('%Y-%m-%d')\n"
        "matrix = data.pivot_table(index='date_label', columns='park_name', values=metric_col, aggfunc='mean')\n"
        "matrix = matrix.sort_index()\n"
        "if metric_col == 'actual_over_pvgis':\n"
        "    plot_df = matrix.clip(lower=0, upper=2) / 2.0\n"
        "    cbar_label = 'Actual/PVGIS intensity, clipped at 2x'\n"
        "else:\n"
        "    max_val = np.nanmax(matrix.to_numpy(dtype=float))\n"
        "    plot_df = matrix / max_val if pd.notna(max_val) and max_val > 0 else matrix\n"
        "    cbar_label = 'Relative intensity'\n"
        "fig_width = max(12, 0.38 * len(plot_df.columns) + 5)\n"
        "fig_height = max(6, min(42, 0.22 * len(plot_df.index) + 2.5))\n"
        "fig, ax = plt.subplots(figsize=(fig_width, fig_height))\n"
        "im = ax.imshow(plot_df.values, aspect='auto', interpolation='nearest', cmap='turbo', vmin=0, vmax=1)\n"
        "title = '" + title_prefix + " - ' + metric_label\n"
        "if pd.notna(from_date) and pd.notna(to_date):\n"
        "    title = f'{title} | {from_date:%Y-%m-%d} to {to_date:%Y-%m-%d}'\n"
        "ax.set_title(title, fontsize=14, fontweight='bold')\n"
        "ax.set_xticks(np.arange(len(plot_df.columns)))\n"
        "ax.set_xticklabels([str(c)[:28] for c in plot_df.columns], rotation=90, fontsize=8)\n"
        "tick_step = max(1, int(np.ceil(len(plot_df.index) / 24)))\n"
        "yticks = np.arange(0, len(plot_df.index), tick_step)\n"
        "ax.set_yticks(yticks)\n"
        "ax.set_yticklabels([plot_df.index[i] for i in yticks], fontsize=8)\n"
        "ax.set_xlabel('Park')\n"
        "ax.set_ylabel('Date')\n"
        "cbar = plt.colorbar(im, ax=ax, fraction=0.025, pad=0.02)\n"
        "cbar.ax.set_ylabel(cbar_label, rotation=90, fontsize=10)\n"
        "ax.grid(False)\n"
        "plt.tight_layout()\n"
        "fig"
    )
    return _to_python_in_excel_formula(code)


def build_workbook() -> Path:
    daily_df = _read_silver_daily_energy(SILVER_CSV)
    pvgis_df = _read_pvgis(PVGIS_CSV)
    metadata_df = _read_park_metadata(PARK_METADATA_CSV)

    daily_out = daily_df[["park_name", "sensor_name", "park_iso_name", "park_id", "date", "value", "day_of_year"]].copy()
    pvgis_out = pvgis_df[["park_name", "park_iso_name", "park_id", "value", "day_of_year"]].copy()

    if "park_id" in metadata_df.columns:
        metadata_out = metadata_df.copy()
    else:
        metadata_out = pd.DataFrame(columns=["park_id", "park_name", "park_iso_name", "price_euro_to_kwh"])

    metadata_park_id_col = (
        get_column_letter(metadata_out.columns.get_loc("park_id") + 1)
        if "park_id" in metadata_out.columns else "J"
    )
    metadata_price_col = (
        get_column_letter(metadata_out.columns.get_loc("price_euro_to_kwh") + 1)
        if "price_euro_to_kwh" in metadata_out.columns else "W"
    )

    park_prices = (
        metadata_out[[c for c in ["park_id", "price_euro_to_kwh"] if c in metadata_out.columns]]
        .dropna(subset=["park_id"]) if "park_id" in metadata_out.columns else pd.DataFrame(columns=["park_id", "price_euro_to_kwh"])
    )
    if not park_prices.empty:
        park_prices = park_prices.drop_duplicates(subset=["park_id"], keep="last")

    default_date = daily_out["date"].max()
    if pd.isna(default_date):
        default_date = pd.Timestamp("today").normalize()

    parks = (
        pvgis_out[["park_id", "park_name"]]
        .dropna(subset=["park_id"])
        .drop_duplicates(subset=["park_id"])
        .sort_values(["park_name", "park_id"])
        .reset_index(drop=True)
    )

    if not park_prices.empty:
        parks = parks.merge(park_prices, on="park_id", how="left")
    else:
        parks["price_euro_to_kwh"] = pd.NA

    default_from_date = daily_out["date"].max() - pd.Timedelta(days=29)
    if pd.isna(default_from_date):
        default_from_date = default_date

    range_detail_out = _build_range_detail_data(daily_out, pvgis_out, parks)
    range_detail_end_row = len(range_detail_out) + 1
    daily_last_row = len(daily_out) + 1
    pvgis_last_row = len(pvgis_out) + 1

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        daily_out.to_excel(writer, sheet_name="Daily Energy", index=False)
        pvgis_out.to_excel(writer, sheet_name="PVGIS Data", index=False)

        workbook = writer.book
        ws_helper_daily = workbook.create_sheet("helper daily analysis")
        _write_helper_daily_analysis_sheet(ws_helper_daily, daily_last_row, pvgis_last_row)

        metadata_out.to_excel(writer, sheet_name="Park Metadata", index=False)
        range_detail_out.to_excel(writer, sheet_name="Range Detail Data", index=False)

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        ws_range_detail_data = workbook["Range Detail Data"]
        ws_range_detail_data.sheet_state = "hidden"
        for col, w in [("A", 12), ("B", 12), ("C", 34), ("D", 42), ("E", 16), ("F", 16), ("G", 18)]:
            ws_range_detail_data.column_dimensions[col].width = w

        ws = workbook.create_sheet("Analysis")

        ws["A1"] = "Selected date"
        ws["B1"] = default_date.to_pydatetime()
        ws["B1"].number_format = "yyyy-mm-dd"

        ws["A2"] = "Day of year"
        ws["B2"] = "=IF(B1=\"\",\"\",B1-DATE(YEAR(B1),1,0))"

        ws["A4"] = "park_id"
        ws["B4"] = "park_name"
        ws["C4"] = "actual_value"
        ws["D4"] = "pvgis_value"
        ws["E4"] = "actual_over_pvgis"

        for cell in ws[4]:
            cell.font = Font(bold=True)

        start_row = 5
        end_row = start_row + len(parks) - 1
        for idx, row in enumerate(parks.itertuples(index=False), start=start_row):
            ws[f"A{idx}"] = row.park_id
            ws[f"B{idx}"] = row.park_name
            ws[f"C{idx}"] = (
                f"=SUMIFS('Daily Energy'!$F:$F,'Daily Energy'!$D:$D,$A{idx},'Daily Energy'!$E:$E,$B$1)"
            )
            ws[f"D{idx}"] = (
                f"=SUMIFS('PVGIS Data'!$D:$D,'PVGIS Data'!$C:$C,$A{idx},'PVGIS Data'!$E:$E,$B$2)"
            )
            ws[f"E{idx}"] = f"=IFERROR(C{idx}/D{idx},NA())"

        ws["G1"] = "Python in Excel (optional)"
        ws["G2"] = "If your Excel supports PY(), this returns day_of_year from B1:"
        ws["G3"] = "=PY(\"import datetime as dt; d=xl('B1'); (d-dt.date(d.year,1,1)).days+1\")"

        chart = BarChart()
        chart.title = "Actual/PVGIS ratio by park for selected date"
        chart.y_axis.title = "actual_over_pvgis"
        chart.x_axis.title = "park"
        chart.height = 10
        chart.width = 20

        data = Reference(ws, min_col=5, min_row=4, max_row=end_row)
        categories = Reference(ws, min_col=2, min_row=5, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "G6")

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["G"].width = 52

        ws_range = workbook.create_sheet("Daily Range Analysis")
        ws_range["A1"] = "From date"
        ws_range["B1"] = default_from_date.to_pydatetime()
        ws_range["B1"].number_format = "yyyy-mm-dd"
        ws_range["A2"] = "To date"
        ws_range["B2"] = default_date.to_pydatetime()
        ws_range["B2"].number_format = "yyyy-mm-dd"

        ws_range["A3"] = "From day_of_year"
        ws_range["B3"] = "=IF(B1=\"\",\"\",B1-DATE(YEAR(B1),1,0))"
        ws_range["A4"] = "To day_of_year"
        ws_range["B4"] = "=IF(B2=\"\",\"\",B2-DATE(YEAR(B2),1,0))"
        ws_range["A5"] = "Metric to plot"
        ws_range["B5"] = "Actual/PVGIS"

        metric_validation = DataValidation(
            type="list",
            formula1='"Actual,PVGIS,Actual/PVGIS"',
            allow_blank=False,
        )
        ws_range.add_data_validation(metric_validation)
        metric_validation.add(ws_range["B5"])

        ws_range["A6"] = "park_id"
        ws_range["B6"] = "park_name"
        ws_range["C6"] = "actual_value_range"
        ws_range["D6"] = "pvgis_value_range"
        ws_range["E6"] = "actual_over_pvgis"
        for cell in ws_range[6]:
            cell.font = Font(bold=True)

        start_row_range = 7
        end_row_range = start_row_range + len(parks) - 1
        for idx, row in enumerate(parks.itertuples(index=False), start=start_row_range):
            ws_range[f"A{idx}"] = row.park_id
            ws_range[f"B{idx}"] = row.park_name
            ws_range[f"C{idx}"] = (
                f"=SUMIFS('Daily Energy'!$F:$F,'Daily Energy'!$D:$D,$A{idx},'Daily Energy'!$E:$E,\">=\"&$B$1,'Daily Energy'!$E:$E,\"<=\"&$B$2)"
            )
            ws_range[f"D{idx}"] = (
                f"=IF($B$3<=$B$4,"
                f"SUMIFS('PVGIS Data'!$D:$D,'PVGIS Data'!$C:$C,$A{idx},'PVGIS Data'!$E:$E,\">=\"&$B$3,'PVGIS Data'!$E:$E,\"<=\"&$B$4),"
                f"SUMIFS('PVGIS Data'!$D:$D,'PVGIS Data'!$C:$C,$A{idx},'PVGIS Data'!$E:$E,\">=\"&$B$3)+"
                f"SUMIFS('PVGIS Data'!$D:$D,'PVGIS Data'!$C:$C,$A{idx},'PVGIS Data'!$E:$E,\"<=\"&$B$4))"
            )
            ws_range[f"E{idx}"] = f"=IFERROR(C{idx}/D{idx},NA())"

        ws_range_chart = BarChart()
        ws_range_chart.title = "Actual/PVGIS ratio by park for selected date range"
        ws_range_chart.y_axis.title = "actual_over_pvgis"
        ws_range_chart.x_axis.title = "park"
        ws_range_chart.height = 10
        ws_range_chart.width = 20
        range_data = Reference(ws_range, min_col=5, min_row=6, max_row=end_row_range)
        range_categories = Reference(ws_range, min_col=2, min_row=7, max_row=end_row_range)
        ws_range_chart.add_data(range_data, titles_from_data=True)
        ws_range_chart.set_categories(range_categories)
        ws_range.add_chart(ws_range_chart, "L6")

        ws_range["G1"] = "Python in Excel heatmap"
        ws_range["G1"].font = Font(bold=True, size=12)
        ws_range["G2"] = _build_python_in_excel_heatmap_formula(
            data_range=f"B6:E{end_row_range}",
            from_date_cell="B1",
            to_date_cell="B2",
            metric_cell="B5",
            title_prefix="Daily Range Analysis Heatmap",
        )

        ws_range.column_dimensions["A"].width = 34
        ws_range.column_dimensions["B"].width = 42
        ws_range.column_dimensions["C"].width = 18
        ws_range.column_dimensions["D"].width = 18
        ws_range.column_dimensions["E"].width = 16
        ws_range.column_dimensions["G"].width = 52
        ws_range.column_dimensions["L"].width = 52

        ws_range_daily = workbook.create_sheet("Range analysis")
        ws_range_daily["A1"] = "From date"
        ws_range_daily["B1"] = default_from_date.to_pydatetime()
        ws_range_daily["B1"].number_format = "yyyy-mm-dd"
        ws_range_daily["A2"] = "To date"
        ws_range_daily["B2"] = default_date.to_pydatetime()
        ws_range_daily["B2"].number_format = "yyyy-mm-dd"
        ws_range_daily["A3"] = "Metric to plot"
        ws_range_daily["B3"] = "Actual/PVGIS"

        daily_metric_validation = DataValidation(
            type="list",
            formula1='"Actual,PVGIS,Actual/PVGIS"',
            allow_blank=False,
        )
        ws_range_daily.add_data_validation(daily_metric_validation)
        daily_metric_validation.add(ws_range_daily["B3"])

        ws_range_daily["A5"] = "Daily rows returned from selected date range"
        ws_range_daily["A6"] = "date"
        ws_range_daily["B6"] = "day_of_year"
        ws_range_daily["C6"] = "park_id"
        ws_range_daily["D6"] = "park_name"
        ws_range_daily["E6"] = "actual_value"
        ws_range_daily["F6"] = "pvgis_value"
        ws_range_daily["G6"] = "actual_over_pvgis"
        for cell in ws_range_daily[6]:
            if cell.value:
                cell.font = Font(bold=True)

        source_data = f"'Range Detail Data'!$A$2:$G${range_detail_end_row}"
        source_dates = f"'Range Detail Data'!$A$2:$A${range_detail_end_row}"
        ws_range_daily["A7"] = (
            f'=FILTER({source_data},({source_dates}>=$B$1)*({source_dates}<=$B$2),"No rows")'
        )
        ws_range_daily["A7"].number_format = "yyyy-mm-dd"

        ws_range_daily["I1"] = "Python in Excel heatmap"
        ws_range_daily["I1"].font = Font(bold=True, size=12)
        ws_range_daily["I2"] = _build_python_in_excel_daily_heatmap_formula(
            data_range=f"'Range Detail Data'!A1:G{range_detail_end_row}",
            from_date_cell="B1",
            to_date_cell="B2",
            metric_cell="B3",
            title_prefix="Range analysis Heatmap",
        )

        ws_range_daily.freeze_panes = "A7"
        for col, w in [("A", 12), ("B", 12), ("C", 34), ("D", 42), ("E", 16), ("F", 16), ("G", 18), ("I", 52)]:
            ws_range_daily.column_dimensions[col].width = w

        ws_month = workbook.create_sheet("Monthly Analysis")
        ws_month["A1"] = "Month (any date inside month)"
        ws_month["B1"] = default_date.to_pydatetime()
        ws_month["B1"].number_format = "yyyy-mm-dd"

        ws_month["A2"] = "Month start"
        ws_month["B2"] = "=DATE(YEAR(B1),MONTH(B1),1)"
        ws_month["A3"] = "Month end"
        ws_month["B3"] = "=EOMONTH(B1,0)"
        ws_month["A4"] = "From day_of_year"
        ws_month["B4"] = "=B2-DATE(YEAR(B2),1,0)"
        ws_month["A5"] = "To day_of_year"
        ws_month["B5"] = "=B3-DATE(YEAR(B3),1,0)"

        # Secondary helpers in cols D-E: last-year-month and YTD bounds
        ws_month["D1"] = "LYM start"
        ws_month["E1"] = "=DATE(YEAR(B1)-1,MONTH(B1),1)"
        ws_month["E1"].number_format = "yyyy-mm-dd"
        ws_month["D2"] = "LYM end"
        ws_month["E2"] = "=EOMONTH(DATE(YEAR(B1)-1,MONTH(B1),1),0)"
        ws_month["E2"].number_format = "yyyy-mm-dd"
        ws_month["D3"] = "YTD start"
        ws_month["E3"] = "=DATE(YEAR(B1),1,1)"
        ws_month["E3"].number_format = "yyyy-mm-dd"
        ws_month["D4"] = "YTD end"
        ws_month["E4"] = "=B3"
        ws_month["E4"].number_format = "yyyy-mm-dd"
        ws_month["D5"] = "LYTD start"
        ws_month["E5"] = "=DATE(YEAR(B1)-1,1,1)"
        ws_month["E5"].number_format = "yyyy-mm-dd"
        ws_month["D6"] = "LYTD end"
        ws_month["E6"] = "=EOMONTH(DATE(YEAR(B1)-1,MONTH(B1),1),0)"
        ws_month["E6"].number_format = "yyyy-mm-dd"

        ws_month["A7"] = "park_id"
        ws_month["B7"] = "park_name"
        ws_month["C7"] = "price_euro_to_kwh"
        ws_month["D7"] = "actual_kwh"
        ws_month["E7"] = "pvgis_kwh"
        ws_month["F7"] = "actual_over_pvgis"
        ws_month["G7"] = "actual_value_eur"
        ws_month["H7"] = "pvgis_value_eur"
        ws_month["I7"] = "lym_kwh"
        ws_month["J7"] = "lym_eur"
        ws_month["K7"] = "month_vs_lym"
        ws_month["L7"] = "ytd_kwh"
        ws_month["M7"] = "ytd_eur"
        ws_month["N7"] = "lytd_kwh"
        ws_month["O7"] = "lytd_eur"
        ws_month["P7"] = "ytd_vs_lytd"
        for cell in ws_month[7]:
            if cell.value:
                cell.font = Font(bold=True)

        start_row_month = 8
        end_row_month = start_row_month + len(parks) - 1
        for idx, row in enumerate(parks.itertuples(index=False), start=start_row_month):
            ws_month[f"A{idx}"] = row.park_id
            ws_month[f"B{idx}"] = row.park_name
            ws_month[f"C{idx}"] = None if pd.isna(row.price_euro_to_kwh) else float(row.price_euro_to_kwh)
            ws_month[f"D{idx}"] = (
                f"=SUMIFS('Daily Energy'!$F:$F,'Daily Energy'!$D:$D,$A{idx},'Daily Energy'!$E:$E,\">=\"&$B$2,'Daily Energy'!$E:$E,\"<=\"&$B$3)"
            )
            ws_month[f"E{idx}"] = (
                f"=IF($B$4<=$B$5,"
                f"SUMIFS('PVGIS Data'!$D:$D,'PVGIS Data'!$C:$C,$A{idx},'PVGIS Data'!$E:$E,\">=\"&$B$4,'PVGIS Data'!$E:$E,\"<=\"&$B$5),"
                f"SUMIFS('PVGIS Data'!$D:$D,'PVGIS Data'!$C:$C,$A{idx},'PVGIS Data'!$E:$E,\">=\"&$B$4)+"
                f"SUMIFS('PVGIS Data'!$D:$D,'PVGIS Data'!$C:$C,$A{idx},'PVGIS Data'!$E:$E,\"<=\"&$B$5))"
            )
            ws_month[f"F{idx}"] = f"=IFERROR(D{idx}/E{idx},NA())"
            ws_month[f"G{idx}"] = f"=IFERROR(D{idx}*C{idx},NA())"
            ws_month[f"H{idx}"] = f"=IFERROR(E{idx}*C{idx},NA())"
            # Last year same month
            _lym = '=SUMIFS(\'Daily Energy\'!$F:$F,\'Daily Energy\'!$D:$D,$A%s,\'Daily Energy\'!$E:$E,">="&$E$1,\'Daily Energy\'!$E:$E,"<="&$E$2)' % idx
            ws_month[f"I{idx}"] = _lym
            ws_month[f"J{idx}"] = f"=IFERROR(I{idx}*C{idx},NA())"
            ws_month[f"K{idx}"] = f"=IFERROR(D{idx}/I{idx},NA())"
            # Year to date (current year)
            _ytd = '=SUMIFS(\'Daily Energy\'!$F:$F,\'Daily Energy\'!$D:$D,$A%s,\'Daily Energy\'!$E:$E,">="&$E$3,\'Daily Energy\'!$E:$E,"<="&$E$4)' % idx
            ws_month[f"L{idx}"] = _ytd
            ws_month[f"M{idx}"] = f"=IFERROR(L{idx}*C{idx},NA())"
            # Last year to date
            _lytd = '=SUMIFS(\'Daily Energy\'!$F:$F,\'Daily Energy\'!$D:$D,$A%s,\'Daily Energy\'!$E:$E,">="&$E$5,\'Daily Energy\'!$E:$E,"<="&$E$6)' % idx
            ws_month[f"N{idx}"] = _lytd
            ws_month[f"O{idx}"] = f"=IFERROR(N{idx}*C{idx},NA())"
            ws_month[f"P{idx}"] = f"=IFERROR(L{idx}/N{idx},NA())"

        ws_month_chart = BarChart()
        ws_month_chart.title = "Monthly monetized actual value by park"
        ws_month_chart.y_axis.title = "EUR"
        ws_month_chart.x_axis.title = "park"
        ws_month_chart.height = 10
        ws_month_chart.width = 20
        month_data = Reference(ws_month, min_col=7, min_row=7, max_row=end_row_month)
        month_categories = Reference(ws_month, min_col=2, min_row=8, max_row=end_row_month)
        ws_month_chart.add_data(month_data, titles_from_data=True)
        ws_month_chart.set_categories(month_categories)
        ws_month.add_chart(ws_month_chart, "R7")

        for col, w in [("A",34),("B",42),("C",18),("D",16),("E",16),("F",16),
                        ("G",16),("H",16),("I",14),("J",14),("K",16),
                        ("L",14),("M",14),("N",14),("O",14),("P",16),("R",52)]:
            ws_month.column_dimensions[col].width = w

        # ── YTD Portfolio History ─────────────────────────────────────────────
        ws_ytd = workbook.create_sheet("YTD Portfolio History")
        ws_ytd["A1"] = "As-of date"
        ws_ytd["B1"] = default_date.to_pydatetime()
        ws_ytd["B1"].number_format = "yyyy-mm-dd"
        ws_ytd["A2"] = "Totals are Jan 1 to this same day/month for each comparison year"

        ws_ytd["A3"] = "Total portfolio YTD (all parks combined)"
        ws_ytd["A1"].font = Font(bold=True, size=12)
        ws_ytd["A3"].font = Font(bold=True)

        HDR_LABELS = ["Year", "YTD Production (kWh)", "YTD Revenue (EUR)", "kWh vs prev year", "EUR vs prev year"]
        HDR_ROW = 5
        for col_idx, label in enumerate(HDR_LABELS, start=1):
            cell = ws_ytd.cell(row=HDR_ROW, column=col_idx, value=label)
            cell.font = Font(bold=True)

        DATA_START = HDR_ROW + 1
        daily_last_row = len(daily_out) + 1
        for r_off in range(10):
            r = DATA_START + r_off
            if r == DATA_START:
                ws_ytd.cell(r, 1).value = "=YEAR($B$1)-9"
            else:
                ws_ytd.cell(r, 1).value = f"=A{r-1}+1"

            end_date_expr = (
                f"DATE(A{r},MONTH($B$1),MIN(DAY($B$1),DAY(EOMONTH(DATE(A{r},MONTH($B$1),1),0))))"
            )

            ws_ytd.cell(r, 2).value = (
                f'=SUMIFS(\'Daily Energy\'!$F$2:$F${daily_last_row},'
                f'\'Daily Energy\'!$E$2:$E${daily_last_row},">="&DATE(A{r},1,1),'
                f'\'Daily Energy\'!$E$2:$E${daily_last_row},"<="&{end_date_expr})'
            )
            ws_ytd.cell(r, 2).number_format = "#,##0"

            ws_ytd.cell(r, 3).value = (
                f'=SUMPRODUCT('
                f'(\'Daily Energy\'!$E$2:$E${daily_last_row}>=DATE(A{r},1,1))*'
                f'(\'Daily Energy\'!$E$2:$E${daily_last_row}<={end_date_expr})*'
                f'\'Daily Energy\'!$F$2:$F${daily_last_row}*'
                f'IFERROR(XLOOKUP(\'Daily Energy\'!$D$2:$D${daily_last_row},'
                f'\'Park Metadata\'!${metadata_park_id_col}:${metadata_park_id_col},'
                f'\'Park Metadata\'!${metadata_price_col}:${metadata_price_col}),0)'
                f')'
            )
            ws_ytd.cell(r, 3).number_format = "#,##0.00"

            if r == DATA_START:
                ws_ytd.cell(r, 4).value = None
                ws_ytd.cell(r, 5).value = None
            else:
                ws_ytd.cell(r, 4).value = f"=IFERROR(B{r}/B{r-1}-1,NA())"
                ws_ytd.cell(r, 4).number_format = "0.00%"
                ws_ytd.cell(r, 5).value = f"=IFERROR(C{r}/C{r-1}-1,NA())"
                ws_ytd.cell(r, 5).number_format = "0.00%"

        data_end = DATA_START + 10 - 1

        # Chart 1 – YTD Production kWh
        chart_kwh = BarChart()
        chart_kwh.title = "Portfolio YTD Production (kWh)"
        chart_kwh.y_axis.title = "kWh"
        chart_kwh.x_axis.title = "Year"
        chart_kwh.height = 12
        chart_kwh.width = 18
        chart_kwh.add_data(Reference(ws_ytd, min_col=2, min_row=HDR_ROW, max_row=data_end), titles_from_data=True)
        chart_kwh.set_categories(Reference(ws_ytd, min_col=1, min_row=DATA_START, max_row=data_end))
        ws_ytd.add_chart(chart_kwh, "G3")

        # Chart 2 – YTD Revenue EUR
        chart_eur = BarChart()
        chart_eur.title = "Portfolio YTD Revenue (EUR)"
        chart_eur.y_axis.title = "EUR"
        chart_eur.x_axis.title = "Year"
        chart_eur.height = 12
        chart_eur.width = 18
        chart_eur.add_data(Reference(ws_ytd, min_col=3, min_row=HDR_ROW, max_row=data_end), titles_from_data=True)
        chart_eur.set_categories(Reference(ws_ytd, min_col=1, min_row=DATA_START, max_row=data_end))
        ws_ytd.add_chart(chart_eur, "Q3")

        # ── Python in Excel chart (Microsoft 365 — Python in Excel) ────────────
        # Replicates visualizations.plot_revenue_by_year exactly:
        # coloured bars vs average, dashed avg reference line, value labels.
        # xl() reads the formula-computed revenue from the data table above.
        _py_range = f"A{DATA_START}:C{data_end}"
        _PY_REVENUE_CODE = (
            "import matplotlib\n"
            "matplotlib.use('Agg')\n"
            "import matplotlib.pyplot as plt\n"
            "data = xl('" + _py_range + "', headers=False)\n"
            "years = [int(x) for x in data.iloc[:, 0]]\n"
            "revenue = data.iloc[:, 2].values.astype(float)\n"
            "avg = float(revenue.mean())\n"
            "colors = ['#27ae60' if v >= 1.10*avg else '#2ecc71' if v >= avg else '#f39c12' if v >= 0.90*avg else '#e74c3c' for v in revenue]\n"
            "fig, ax = plt.subplots(figsize=(12, 7), facecolor='white')\n"
            "ax.bar(range(len(years)), revenue, color=colors, alpha=0.85, edgecolor='#34495e', linewidth=1.5, width=0.6)\n"
            "for i, (yr, val) in enumerate(zip(years, revenue)):\n"
            "    lbl = '{:,.0f}'.format(val) + chr(10) + 'EUR'\n"
            "    ax.text(i, val + max(revenue)*0.02, lbl, ha='center', va='bottom', fontsize=10, fontweight='bold', bbox=dict(boxstyle='round,pad=0.3', facecolor='yellow', alpha=0.3, edgecolor='none'))\n"
            "ax.axhline(avg, color='#e74c3c', linestyle='--', linewidth=2, alpha=0.8, label='Average: {:,.0f} EUR'.format(avg))\n"
            "ax.set_xticks(range(len(years)))\n"
            "ax.set_xticklabels([str(y) for y in years], fontsize=11, fontweight='bold')\n"
            "ax.set_ylabel('Revenue [EUR]', fontsize=12, fontweight='bold', color='#34495e')\n"
            "ax.set_title('Portfolio YTD Revenue by Year', fontsize=14, fontweight='bold', color='#2c3e50', pad=20)\n"
            "ax.grid(axis='y', alpha=0.3, linestyle='--', linewidth=1)\n"
            "ax.set_facecolor('#f8f9fa')\n"
            "ax.legend(fontsize=10, loc='upper left', frameon=True, shadow=True, fancybox=True)\n"
            "ax.spines['top'].set_visible(False)\n"
            "ax.spines['right'].set_visible(False)\n"
            "ax.spines['left'].set_color('#34495e')\n"
            "ax.spines['bottom'].set_color('#34495e')\n"
            "plt.tight_layout()\n"
            "fig"
        )
        _py_label_row = data_end + 3
        _py_chart_row = _py_label_row + 1
        ws_ytd.cell(_py_label_row, 1).value = (
            "Python in Excel chart — requires Microsoft 365 with Python in Excel enabled "
            "(reads the YTD Revenue column above; recalculates when B1 changes):"
        )
        ws_ytd.cell(_py_label_row, 1).font = Font(bold=True)
        ws_ytd.cell(_py_chart_row, 1).value = '=PY("' + _PY_REVENUE_CODE + '")'

        for col, w in [("A", 52), ("B", 24), ("C", 20), ("D", 18), ("E", 18)]:
            ws_ytd.column_dimensions[col].width = w

    return OUT_XLSX


if __name__ == "__main__":
    out_path = build_workbook()
    print(f"Workbook created: {out_path}")
